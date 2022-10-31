import requests
import openpyxl
import json
import math
import time
import unidecode
import threading

max_request_retries = 3
retry_failure_delay = 3

nhl_schedule_url_format = "https://statsapi.web.nhl.com/api/v1/schedule?startDate={}&endDate={}&gameType=P,R"

request_headers = {
    "User-Agent" : "NHLCompareRedditBot"
}

request_headers= {})

def main():
    outdoor_games = openpyxl.load_workbook("outdoor_games.xlsx")

    outdoor_games_arr = []

    sheet_obj = outdoor_games["Sheet1"]
    n_rows = sheet_obj.max_row + 1
    n_columns = sheet_obj.max_column + 1
    for r in range(2, n_rows):
        date = sheet_obj.cell(r, 1).value
        name = sheet_obj.cell(r, 2).value
        away = sheet_obj.cell(r, 3).value
        home = sheet_obj.cell(r, 4).value

        outdoor_games_arr.append({
            "date" : date.date() if date else None,
            "name" : name,
            "away" : away,
            "home" : home
        })

    print(len(outdoor_games_arr))

    outdoor_game_ids = {}
    for outdoor_game in outdoor_games_arr:
        if outdoor_game["date"]:
            found_match = False
            scheudle_url = nhl_schedule_url_format.format(outdoor_game["date"], outdoor_game["date"])
            response = url_request(scheudle_url)
            data = json.load(response)

            for date in data["dates"]:
                for game in date["games"]:
                    if unidecode.unidecode(game["teams"]["away"]["team"]["name"]) == outdoor_game["away"] and unidecode.unidecode(game["teams"]["home"]["team"]["name"]) == outdoor_game["home"]:
                        outdoor_game_ids[game["gamePk"]] = outdoor_game["name"]
                        found_match = True
                        break
    
    print(len(outdoor_game_ids))

    with open("outdoor_games.json", "w") as file:
        file.write(json.dumps(outdoor_game_ids, indent=4, sort_keys=True))

def url_request(url, timeout=30):
    failed_counter = 0
    while(True):
        try:
            response = requests.get(url, timeout=timeout, headers=request_headers)
            response.raise_for_status()
            return response.content
        except requests.exceptions.HTTPError as err:
            if err.response.status_code == 403:
                raise
            else:
                failed_counter += 1
                if failed_counter > max_request_retries:
                    raise
        except Exception:
            failed_counter += 1
            if failed_counter > max_request_retries:
                raise

        delay_step = 10
        print("#" + str(threading.get_ident()) + "#   " + "Retrying in " + str(retry_failure_delay) + " seconds to allow request to " + url + " to chill")
        time_to_wait = int(math.ceil(float(retry_failure_delay)/float(delay_step)))
        for i in range(retry_failure_delay, 0, -time_to_wait):
            print("#" + str(threading.get_ident()) + "#   " + str(i))
            time.sleep(time_to_wait)
        print("#" + str(threading.get_ident()) + "#   " + "0")

if __name__ == "__main__":
    main()